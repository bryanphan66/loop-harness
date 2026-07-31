#!/usr/bin/env python3
"""Excel map du lieu cong viec GitHub Issue va Plane Work Item.
v3 (260727): chu viet plain, de hieu nhu go tay; bo het ky tu kieu AI (cham giua,
gach ngang dai, mui ten); dung dau phay va cau doi thuong. Tieng Viet co dau."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/trung/Desktop/Workspace/videcode-harness/plans/github-plane-field-mapping-sync-260724.xlsx"

# (loai, gia tri, github, plane, sync, ghi chu) — chu plain, khong ky tu AI
ROWS = [
  ("Khóa nối hai bên (mã liên kết)",
   "Số issue, ví dụ 155, 213 (mọi issue). Mã [F-NNN] chỉ có ở issue feature, ví dụ [F-046]",
   "number = số issue (mã nối chính). Riêng issue feature còn có [F-NNN] ở đầu tiêu đề; issue bug/enhancement KHÔNG có [F-NNN]",
   "external_id, chính là số issue GitHub",
   "Có (chức năng có sẵn)",
   "Plane nối hai bên bằng SỐ issue (external_id), áp cho mọi issue. [F-NNN] chỉ là mã phụ bền do người đặt trong tiêu đề feature, không phải cái Plane dùng để nối."),
  ("Tiêu đề",
   "Chữ tự do",
   "title", "name", "Có (chức năng có sẵn)",
   "Tự đồng bộ trong vài giây."),
  ("Mô tả",
   "Chữ, có định dạng Markdown",
   "trường .body (Markdown)", "trường description_html (HTML)",
   "Một phần",
   "Chức năng có sẵn chỉ đổ chữ thô. Muốn hiển thị đẹp như in đậm, tiêu đề thì phải viết thêm tính năng trong Plane."),
  ("Trạng thái công việc",
   "Backlog, In Progress, Dev Done, QC, In Review, Done, Cancelled. Tổng 7 trạng thái.",
   "trường tùy chỉnh tên States",
   "trường state, có 7 trạng thái trùng tên",
   "Chưa",
   "Chức năng có sẵn chỉ đồng bộ đóng hay mở. Trạng thái chi tiết nằm ở trường tùy chỉnh, phải viết thêm tính năng trong Plane."),
  ("Đóng hay mở",
   "open, closed",
   "trạng thái open hoặc closed",
   "open thành Backlog, closed thành Done, theo cấu hình",
   "Có (chức năng có sẵn)",
   "Đồng bộ ngay theo cấu hình trạng thái trên giao diện."),
  ("Nhóm chức năng (Module)",
   "Authentication, User Management & RBAC, Course Management, Learning & Certificates, "
   "Commerce & Payments, Marketing & Automation, CRM, Analytics & Reports, Website Builder & CMS, "
   "Blog, Helpdesk, Notifications, Platform & Integrations. Tổng 13 nhóm.",
   "trường tùy chỉnh tên Module",
   "gán vào Module, quan hệ module-issues",
   "Chưa",
   "Chức năng có sẵn không đọc trường tùy chỉnh, phải viết thêm tính năng trong Plane."),
  ("Giai đoạn (Phase)",
   "Phase 1, Phase 1.5, Phase 2, Phase 3",
   "Milestone",
   "Cycle, có 4 cái trùng tên",
   "Chưa",
   "Đưa Milestone bên GitHub thành Cycle bên Plane. Chức năng có sẵn chưa làm, phải viết thêm."),
  ("Loại công việc (Type)",
   "Feature, Bug, Enhancement",
   "Issue Type",
   "label trùng tên",
   "Chưa",
   "Đưa Issue Type thành nhãn bên Plane. Chức năng có sẵn chưa làm, phải viết thêm."),
  ("Mức ưu tiên (Priority)",
   "Urgent, High, Medium, Low, hoặc để trống",
   "trường tùy chỉnh tên Priority",
   "trường priority có sẵn",
   "Chưa",
   "Đưa mức ưu tiên GitHub thành mức ưu tiên Plane. Chức năng có sẵn chưa làm, phải viết thêm."),
  ("Người phụ trách (Assignee)",
   "Tài khoản thành viên đã ghép",
   "assignees, lấy login",
   "assignees, lưu theo UUID",
   "Có (chức năng có sẵn)",
   "Đồng bộ ngay, thông qua bảng ghép thành viên trên giao diện."),
  ("Nhãn (Labels)",
   "feature, bug, Enhancement, Plane, Github",
   "label",
   "label",
   "Có (chức năng có sẵn)",
   "Nhãn bên GitHub thành nhãn bên Plane, đồng bộ ngay."),
  ("Việc cha, việc con (Parent)",
   "Số của issue cha, ví dụ 163",
   "trường parent, và sub_issues",
   "trường parent",
   "Chưa",
   "Nối việc con lên việc cha. Chức năng có sẵn chưa làm, phải viết thêm."),
  ("Phụ thuộc, chặn nhau (Dependencies)",
   "Quan hệ chặn, kiểu blocks hoặc blocked by",
   "issue_dependencies",
   "issue-relations",
   "Không",
   "Chỉ có trên GitHub, không đồng bộ."),
  ("Bình luận (Comments)",
   "Chữ bình luận",
   "comments",
   "comments",
   "Có (chức năng có sẵn)",
   "Bình luận đồng bộ ngay."),
  ("Tệp đính kèm (Attachments)",
   "Ảnh hoặc tệp trong nội dung, bình luận",
   "dán vào nội dung hoặc bình luận",
   "issue-attachments",
   "Không",
   "Chức năng có sẵn không đồng bộ tệp."),
  ("Liên kết mã nguồn (PR, commit)",
   "Link Pull Request hoặc commit",
   "khung Development, link Pull Request",
   "gắn qua chức năng có sẵn",
   "Có (chức năng có sẵn)",
   "Có Pull Request thì Plane tự gắn link, đồng bộ ngay."),
]

wb = Workbook()
BRAND = PatternFill("solid", fgColor="24292F")
HEAD = PatternFill("solid", fgColor="255986")
GREEN = PatternFill("solid", fgColor="C7ECD6")
ORA = PatternFill("solid", fgColor="FBE0C4")
RED = PatternFill("solid", fgColor="F5D0D0")
ZEBRA = PatternFill("solid", fgColor="F5F7FA")
white = Font(color="FFFFFF", bold=True)
bold = Font(bold=True)
thin = Side(style="thin", color="D0D5DD")
border = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws = wb.active
ws.title = "Map GitHub-Plane"
ws.merge_cells("A1:G1")
c = ws["A1"]; c.value = "Bản đồ dữ liệu công việc: GitHub Issue và Plane Work Item (kèm giá trị và có đồng bộ được không)"
c.font = Font(color="FFFFFF", bold=True, size=13); c.fill = BRAND
c.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 26
ws.merge_cells("A2:G2")
ws["A2"] = ("Cách làm mong muốn: chỉ dùng chức năng Repository & project sync có sẵn của Plane, cấu hình bằng tay trên "
            "giao diện. Không dùng cron, không script, không API chạy ngầm. Dòng ghi Chưa nghĩa là chức năng có sẵn "
            "chưa làm được, cần viết thêm tính năng trong Plane, vì Plane là mã nguồn mở và mình tự cài. Cập nhật ngày 27 tháng 07 năm 2026.")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="center"); ws.row_dimensions[2].height = 40

headers = ["STT", "Loại dữ liệu", "Giá trị hiện có", "Trên GitHub Issue", "Trên Plane Work Item", "Có đồng bộ được không", "Giải thích và ghi chú"]
hr = 3
for j, h in enumerate(headers, 1):
    cell = ws.cell(hr, j, h); cell.fill = HEAD; cell.font = white; cell.alignment = center; cell.border = border
for j, w in enumerate([5, 24, 42, 28, 28, 18, 46], 1):
    ws.column_dimensions[get_column_letter(j)].width = w

def sync_fill(v):
    return {"Có (chức năng có sẵn)": GREEN, "Một phần": ORA, "Chưa": RED, "Không": RED}.get(v)

# Pham vi tren GitHub: Org = org Issue Fields (States/Module/Priority) + Issue Type; con lai = Repo.
# Thu tu khop ROWS: 1 khoa noi,2 tieu de,3 mo ta,4 States,5 dong/mo,6 Module,7 Milestone,
# 8 Type,9 Priority,10 assignee,11 labels,12 parent,13 deps,14 comment,15 attach,16 PR.
SCOPE = ["Repo","Repo","Repo","Org","Repo","Org","Repo","Org","Org","Repo","Repo","Repo","Repo","Repo","Repo","Repo"]

r = hr + 1
for i, (loai, val, gh, pl, sync, note) in enumerate(ROWS, 1):
    ws.cell(r, 1, i); ws.cell(r, 2, loai); ws.cell(r, 3, val)
    ws.cell(r, 4, f"{gh}  (cấp {SCOPE[i-1]})"); ws.cell(r, 5, pl); ws.cell(r, 6, sync); ws.cell(r, 7, note)
    zeb = ZEBRA if i % 2 == 0 else None
    for j in range(1, 8):
        cc = ws.cell(r, j); cc.border = border
        cc.alignment = center if j in (1, 6) else wrap
        if j == 2:
            cc.font = bold
        if j == 6:
            f = sync_fill(sync); cc.fill = f if f else PatternFill()
        elif zeb:
            cc.fill = zeb
    r += 1

r += 1
ws.merge_cells(f"A{r}:G{r}")
ws.cell(r, 1, "Giải thích cột Có đồng bộ được không. "
              "Có (chức năng có sẵn) là Plane tự đồng bộ ngay, chỉ cần bật và cấu hình trên giao diện. "
              "Một phần là làm được nhưng chưa trọn. "
              "Chưa là chức năng có sẵn chưa làm được, phải viết thêm tính năng trong Plane. "
              "Không là không đồng bộ, dữ liệu chỉ có trên GitHub. Đã bỏ hẳn cron và script.")
ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="center"); ws.cell(r, 1).font = Font(italic=True, size=10)
ws.row_dimensions[r].height = 44
r += 1
for k, (lab, fill) in enumerate([("Có (chức năng có sẵn)", GREEN), ("Một phần", ORA), ("Chưa, phải viết thêm", RED)]):
    cc = ws.cell(r, 2 + k, lab); cc.fill = fill; cc.border = border; cc.alignment = center; cc.font = bold

ws.freeze_panes = "A4"
ws.auto_filter.ref = f"A{hr}:G{hr + len(ROWS)}"
wb.save(OUT)
print("saved:", OUT, "| rows:", len(ROWS))
